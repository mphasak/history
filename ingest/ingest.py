"""
ingest.py — Human History Simulator

Reads an extraction template (template_v0.3.xlsx) and upserts its contents into
Postgres + PostGIS, in dependency order, with referential-integrity validation
performed in-memory before any database writes.

Usage:
    python ingest.py path/to/template_v0.3.xlsx \\
        --dsn 'postgresql://user:pass@localhost:5432/history_sim' \\
        [--dry-run]

Idempotent: re-running with the same spreadsheet is a no-op for unchanged rows.

Design notes
------------
- Validation is split from writes. We load every sheet into memory, cross-check
  references, and only open a database transaction if validation passes. This
  keeps the database clean when the spreadsheet has typos.
- Dependency order matters: Sources -> Geographies -> Traits -> TraitRelations
  -> Carriers -> CarrierTraitMix -> TraitObservations -> PhysicalFeatures
  -> PaleoclimateStates -> PropagationEvents -> Claims -> Perspectives
  -> PerspectiveEndorsements.
- Example rows (green-fill in the spreadsheet) are documentation, but for v0.3
  Phase 0 we ingest them as real data because they ARE the seed dataset.
- Empty cells -> NULL. Whitespace-only cells -> NULL. We never insert empty strings.
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass, field
from typing import Any, Iterable

from openpyxl import load_workbook          # openpyxl>=3.1
# psycopg is imported lazily inside upsert() so dry-run works without libpq.


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _norm(v: Any) -> Any:
    """Normalize a cell value: strip strings, treat blank as None."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def _bool(v: Any) -> bool | None:
    v = _norm(v)
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.upper() == 'TRUE'
    return bool(v)


def _int(v: Any) -> int | None:
    v = _norm(v)
    if v is None:
        return None
    return int(v)


def _float(v: Any) -> float | None:
    v = _norm(v)
    if v is None:
        return None
    return float(v)


def _split_ids(v: Any) -> list[str]:
    """Source IDs are stored as semicolon-separated strings; split and strip."""
    v = _norm(v)
    if v is None:
        return []
    return [p.strip() for p in str(v).split(';') if p.strip()]


def _authors(v: Any) -> list[str]:
    """Author strings stored as semicolon-separated; return as list for text[]."""
    v = _norm(v)
    if v is None:
        return []
    return [a.strip() for a in str(v).split(';') if a.strip()]


def _point_wkt(lat: float | None, lon: float | None) -> str | None:
    """Build a PostGIS POINT WKT in WGS84. Returns None if either coord missing."""
    if lat is None or lon is None:
        return None
    return f'SRID=4326;POINT({lon} {lat})'


def _read_rows(ws, headers: list[str]) -> Iterable[dict]:
    """Yield rows as dicts keyed by header. Skips fully-empty rows."""
    actual = [c.value for c in ws[1]]
    if [h.strip() if isinstance(h, str) else h for h in actual[:len(headers)]] != headers:
        raise ValueError(
            f'Sheet {ws.title!r} header mismatch.\n'
            f'  expected: {headers}\n'
            f'  actual  : {actual[:len(headers)]}'
        )
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        yield {h: row[i] if i < len(row) else None for i, h in enumerate(headers)}


# ---------------------------------------------------------------------------
# Loaded dataset
# ---------------------------------------------------------------------------

@dataclass
class Dataset:
    sources: list[dict] = field(default_factory=list)
    geographies: list[dict] = field(default_factory=list)
    traits: list[dict] = field(default_factory=list)
    trait_relations: list[dict] = field(default_factory=list)
    carriers: list[dict] = field(default_factory=list)
    carrier_trait_mix: list[dict] = field(default_factory=list)
    trait_observations: list[dict] = field(default_factory=list)
    physical_features: list[dict] = field(default_factory=list)
    paleoclimate_states: list[dict] = field(default_factory=list)
    propagation_events: list[dict] = field(default_factory=list)
    claims: list[dict] = field(default_factory=list)
    perspectives: list[dict] = field(default_factory=list)
    perspective_endorsements: list[dict] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Sheet loaders (one per sheet — header lists must match build_v3.py exactly)
# ---------------------------------------------------------------------------

def load(path: str) -> Dataset:
    wb = load_workbook(path, data_only=True)
    ds = Dataset()

    # Sources
    ws = wb['Sources']
    headers = ['Source ID', 'Type', 'Citation', 'Author(s)', 'Year', 'Pages / Section',
               'DOI / URL', 'Edition', 'Default Confidence Weight (0-1)',
               'Superseded By', 'Tags', 'Notes']
    for r in _read_rows(ws, headers):
        ds.sources.append({
            'id': _norm(r['Source ID']),
            'type': _norm(r['Type']),
            'citation': _norm(r['Citation']),
            'authors': _authors(r['Author(s)']),       # fixed: list not string
            'year': _int(r['Year']),
            'pages_section': _norm(r['Pages / Section']),
            'doi_url': _norm(r['DOI / URL']),
            'edition': _norm(r['Edition']),
            'default_weight': _float(r['Default Confidence Weight (0-1)']) or 1.0,
            'superseded_by': _norm(r['Superseded By']),
            'tags': _split_ids(r['Tags']),
        })

    # Geographies
    ws = wb['Geographies']
    headers = ['Geo ID', 'Display Name', 'Type', 'Center Lat', 'Center Lon',
               'Bounding Box (W,S,E,N)', 'Parent Geo ID', 'Notes']
    for r in _read_rows(ws, headers):
        ds.geographies.append({
            'id': _norm(r['Geo ID']),
            'display_name': _norm(r['Display Name']),
            'type': _norm(r['Type']),
            'lat': _float(r['Center Lat']),
            'lon': _float(r['Center Lon']),
            'bbox': _norm(r['Bounding Box (W,S,E,N)']),
            'parent_id': _norm(r['Parent Geo ID']),
        })

    # Traits
    ws = wb['Traits']
    headers = ['Trait ID', 'Domain', 'Display Name', 'Aliases', 'First Detected (year)',
               'Last Detected (year)', 'Origin Region', 'Origin Lat', 'Origin Lon',
               'Defining Sample(s) / Exemplars', 'Description', 'Source IDs',
               'Confidence (1-5)', 'Politically Sensitive?', 'Notes']
    for r in _read_rows(ws, headers):
        ds.traits.append({
            'id': _norm(r['Trait ID']),
            'domain': _norm(r['Domain']),
            'display_name': _norm(r['Display Name']),
            'aliases': _split_ids(r['Aliases']),
            'first_year': _int(r['First Detected (year)']),
            'last_year': _int(r['Last Detected (year)']),
            'origin_region': _norm(r['Origin Region']),
            'lat': _float(r['Origin Lat']),
            'lon': _float(r['Origin Lon']),
            'exemplars': _split_ids(r['Defining Sample(s) / Exemplars']),
            'description': _norm(r['Description']),
            'source_ids': _split_ids(r['Source IDs']),
            'confidence': _int(r['Confidence (1-5)']),
            'politically_sensitive': _bool(r['Politically Sensitive?']) or False,
        })

    # TraitRelations
    ws = wb['TraitRelations']
    headers = ['Relation ID', 'Child Trait ID', 'Parent Trait ID', 'Relation Type',
               'Weight', 'Source IDs', 'Confidence (1-5)', 'Notes']
    for r in _read_rows(ws, headers):
        ds.trait_relations.append({
            'id': _norm(r['Relation ID']),
            'child_id': _norm(r['Child Trait ID']),
            'parent_id': _norm(r['Parent Trait ID']),
            'relation_type': _norm(r['Relation Type']),
            'weight': _float(r['Weight']),
            'source_ids': _split_ids(r['Source IDs']),
            'confidence': _int(r['Confidence (1-5)']),
        })

    # Carriers
    ws = wb['Carriers']
    headers = ['Carrier ID', 'Display Name', 'Type', 'Period (Display)',
               'Date Min (year)', 'Date Max (year)', 'Region', 'Center Lat', 'Center Lon',
               'Bounding Box (W,S,E,N)', 'Archaeological Culture', 'Linguistic Affiliation',
               'Source IDs', 'Confidence (1-5)', 'Notes']
    for r in _read_rows(ws, headers):
        ds.carriers.append({
            'id': _norm(r['Carrier ID']),
            'display_name': _norm(r['Display Name']),
            'type': _norm(r['Type']),
            'date_min': _int(r['Date Min (year)']),
            'date_max': _int(r['Date Max (year)']),
            'region': _norm(r['Region']),
            'lat': _float(r['Center Lat']),
            'lon': _float(r['Center Lon']),
            'bbox': _norm(r['Bounding Box (W,S,E,N)']),
            'arch_culture': _norm(r['Archaeological Culture']),
            'linguistic': _norm(r['Linguistic Affiliation']),
            'source_ids': _split_ids(r['Source IDs']),
            'confidence': _int(r['Confidence (1-5)']),
        })

    # CarrierTraitMix
    ws = wb['CarrierTraitMix']
    headers = ['Mix ID', 'Carrier ID', 'As-of Year', 'Domain', 'Trait ID', 'Fraction',
               'Stderr', 'Source IDs', 'Confidence (1-5)', 'Notes']
    for r in _read_rows(ws, headers):
        ds.carrier_trait_mix.append({
            'id': _norm(r['Mix ID']),
            'carrier_id': _norm(r['Carrier ID']),
            'as_of_year': _int(r['As-of Year']),
            'domain': _norm(r['Domain']),
            'trait_id': _norm(r['Trait ID']),
            'fraction': _float(r['Fraction']),
            'stderr': _float(r['Stderr']),
            'source_ids': _split_ids(r['Source IDs']),
            'confidence': _int(r['Confidence (1-5)']),
        })

    # TraitObservations
    ws = wb['TraitObservations']
    headers = ['Obs ID', 'Carrier ID', 'Sample / Site / Document', 'Date Min', 'Date Max',
               'Lat', 'Lon', 'Domain', 'Trait ID', 'Estimated Fraction', 'Stderr',
               'Method', 'Source IDs', 'Confidence (1-5)', 'Notes']
    for r in _read_rows(ws, headers):
        ds.trait_observations.append({
            'id': _norm(r['Obs ID']),
            'carrier_id': _norm(r['Carrier ID']),
            'sample': _norm(r['Sample / Site / Document']),
            'date_min': _int(r['Date Min']),
            'date_max': _int(r['Date Max']),
            'lat': _float(r['Lat']),
            'lon': _float(r['Lon']),
            'domain': _norm(r['Domain']),
            'trait_id': _norm(r['Trait ID']),
            'fraction': _float(r['Estimated Fraction']),
            'stderr': _float(r['Stderr']),
            'method': _norm(r['Method']),
            'source_ids': _split_ids(r['Source IDs']),
            'confidence': _int(r['Confidence (1-5)']),
        })

    # PhysicalFeatures
    ws = wb['PhysicalFeatures']
    headers = ['Feature ID', 'Type', 'Display Name', 'As-of Year', 'Geometry Description',
               'Center Lat', 'Center Lon', 'Bounding Box (W,S,E,N)', 'Source IDs',
               'Confidence (1-5)', 'Notes']
    for r in _read_rows(ws, headers):
        ds.physical_features.append({
            'id': _norm(r['Feature ID']),
            'type': _norm(r['Type']),
            'display_name': _norm(r['Display Name']),
            'as_of_year': _int(r['As-of Year']),
            'geometry_desc': _norm(r['Geometry Description']),
            'lat': _float(r['Center Lat']),
            'lon': _float(r['Center Lon']),
            'bbox': _norm(r['Bounding Box (W,S,E,N)']),
            'source_ids': _split_ids(r['Source IDs']),
            'confidence': _int(r['Confidence (1-5)']),
        })

    # PaleoclimateStates
    ws = wb['PaleoclimateStates']
    headers = ['State ID', 'Year', 'Scope', 'Region', 'Sea Level vs Present (m)',
               'Mean Temp Anomaly (C)', 'Ice Volume (relative)', 'CO2 (ppm)',
               'Source IDs', 'Confidence (1-5)', 'Notes']
    for r in _read_rows(ws, headers):
        ds.paleoclimate_states.append({
            'id': _norm(r['State ID']),
            'year': _int(r['Year']),
            'scope': _norm(r['Scope']),
            'region': _norm(r['Region']),
            'sea_level': _float(r['Sea Level vs Present (m)']),
            'temp_anomaly': _float(r['Mean Temp Anomaly (C)']),
            'ice_volume': _norm(r['Ice Volume (relative)']),
            'co2_ppm': _float(r['CO2 (ppm)']),
            'source_ids': _split_ids(r['Source IDs']),
            'confidence': _int(r['Confidence (1-5)']),
        })

    # PropagationEvents
    ws = wb['PropagationEvents']
    headers = ['Event ID', 'Display Name', 'Domain', 'Date Min', 'Date Max',
               'Source Trait ID(s)', 'Source Region', 'Source Lat', 'Source Lon',
               'Destination Region', 'Dest Lat', 'Dest Lon', 'Resulting Trait Mix Change',
               'Mechanism', 'Source IDs', 'Confidence (1-5)', 'Politically Sensitive?',
               'Notes']
    for r in _read_rows(ws, headers):
        ds.propagation_events.append({
            'id': _norm(r['Event ID']),
            'display_name': _norm(r['Display Name']),
            'domain': _norm(r['Domain']),
            'date_min': _int(r['Date Min']),
            'date_max': _int(r['Date Max']),
            'source_trait_ids': _split_ids(r['Source Trait ID(s)']),
            'source_region': _norm(r['Source Region']),
            'source_lat': _float(r['Source Lat']),
            'source_lon': _float(r['Source Lon']),
            'dest_region': _norm(r['Destination Region']),
            'dest_lat': _float(r['Dest Lat']),
            'dest_lon': _float(r['Dest Lon']),
            'mix_change': _norm(r['Resulting Trait Mix Change']),
            'mechanism': _norm(r['Mechanism']),
            'source_ids': _split_ids(r['Source IDs']),
            'confidence': _int(r['Confidence (1-5)']),
            'politically_sensitive': _bool(r['Politically Sensitive?']) or False,
        })

    # Claims
    ws = wb['Claims']
    headers = ['Claim ID', 'Subject Type', 'Subject ID', 'Statement',
               'Quantitative Value (optional)', 'Default Supporting Source IDs',
               'Default Disputing Source IDs', 'Aggregated Confidence (1-5)',
               'Politically Sensitive?', 'Notes']
    for r in _read_rows(ws, headers):
        ds.claims.append({
            'id': _norm(r['Claim ID']),
            'subject_type': _norm(r['Subject Type']),
            'subject_id': _norm(r['Subject ID']),
            'statement': _norm(r['Statement']),
            'quant_value': _norm(r['Quantitative Value (optional)']),
            'supporting_ids': _split_ids(r['Default Supporting Source IDs']),
            'disputing_ids': _split_ids(r['Default Disputing Source IDs']),
            'confidence': _int(r['Aggregated Confidence (1-5)']),
            'politically_sensitive': _bool(r['Politically Sensitive?']) or False,
        })

    # Perspectives
    ws = wb['Perspectives']
    headers = ['Perspective ID', 'Display Name', 'Domain Scope', 'Summary',
               'Proponents / Tradition', 'Methodology Notes', 'Parent Perspective',
               'Default Active?', 'Status', 'Notes']
    for r in _read_rows(ws, headers):
        scope_raw = _norm(r['Domain Scope'])
        scope = []
        if scope_raw and scope_raw.lower() != 'all':
            scope = [s.strip() for s in scope_raw.replace(';', ',').split(',') if s.strip()]
        ds.perspectives.append({
            'id': _norm(r['Perspective ID']),
            'display_name': _norm(r['Display Name']),
            'domain_scope': scope,
            'summary': _norm(r['Summary']),
            'proponents': _norm(r['Proponents / Tradition']),
            'methodology_notes': _norm(r['Methodology Notes']),
            'parent_id': _norm(r['Parent Perspective']),
            'default_active': _bool(r['Default Active?']) or False,
            'status': _norm(r['Status']) or 'provisional',
        })

    # PerspectiveEndorsements
    ws = wb['PerspectiveEndorsements']
    headers = ['Endorsement ID', 'Perspective ID', 'Subject Type', 'Subject ID',
               'Stance', 'Override Statement', 'Override Quantitative Value',
               'Source Weight Overrides', 'Notes']
    for r in _read_rows(ws, headers):
        ds.perspective_endorsements.append({
            'id': _norm(r['Endorsement ID']),
            'perspective_id': _norm(r['Perspective ID']),
            'subject_type': _norm(r['Subject Type']),
            'subject_id': _norm(r['Subject ID']),
            'stance': _norm(r['Stance']),
            'override_statement': _norm(r['Override Statement']),
            'override_quant': _norm(r['Override Quantitative Value']),
            'source_weight_overrides': _norm(r['Source Weight Overrides']),
        })

    return ds


# ---------------------------------------------------------------------------
# Validation (in-memory, before any writes)
# ---------------------------------------------------------------------------

def validate(ds: Dataset) -> list[str]:
    """Return list of error strings. Empty list = OK."""
    errors: list[str] = []

    # Build ID indexes
    src_ids = {s['id'] for s in ds.sources if s['id']}
    geo_ids = {g['id'] for g in ds.geographies if g['id']}
    trait_ids = {t['id'] for t in ds.traits if t['id']}
    carrier_ids = {c['id'] for c in ds.carriers if c['id']}
    persp_ids = {p['id'] for p in ds.perspectives if p['id']}
    claim_ids = {c['id'] for c in ds.claims if c['id']}

    DOMAINS = {'genetic', 'technological', 'linguistic', 'religious', 'ideological',
               'artistic', 'institutional', 'material_culture', 'other'}

    def chk_sources(label: str, row_id: str, ids: list[str]) -> None:
        for sid in ids:
            if sid not in src_ids:
                errors.append(f'{label} {row_id!r}: unknown Source ID {sid!r}')

    # Sources: unique IDs, weights in [0, 1]
    seen = set()
    for s in ds.sources:
        if not s['id']:
            errors.append(f'Source row missing ID: {s!r}')
            continue
        if s['id'] in seen:
            errors.append(f'Duplicate Source ID: {s["id"]!r}')
        seen.add(s['id'])
        if s['default_weight'] is not None and not (0 <= s['default_weight'] <= 1):
            errors.append(f'Source {s["id"]}: default_weight {s["default_weight"]} out of [0, 1]')
        if s['superseded_by'] and s['superseded_by'] not in src_ids:
            errors.append(f'Source {s["id"]}: superseded_by unknown source {s["superseded_by"]!r}')

    # Geographies: parent_id must resolve
    for g in ds.geographies:
        if g['parent_id'] and g['parent_id'] not in geo_ids:
            errors.append(f'Geography {g["id"]}: unknown parent_id {g["parent_id"]!r}')

    # Traits
    for t in ds.traits:
        if not t['id']:
            errors.append(f'Trait row missing ID: {t!r}'); continue
        if t['domain'] not in DOMAINS:
            errors.append(f'Trait {t["id"]}: invalid domain {t["domain"]!r}')
        chk_sources('Trait', t['id'], t['source_ids'])

    # TraitRelations: child & parent must exist
    for r in ds.trait_relations:
        if r['child_id'] not in trait_ids:
            errors.append(f'TraitRelation {r["id"]}: unknown child {r["child_id"]!r}')
        if r['parent_id'] not in trait_ids:
            errors.append(f'TraitRelation {r["id"]}: unknown parent {r["parent_id"]!r}')
        chk_sources('TraitRelation', r['id'], r['source_ids'])

    # Carriers: date_min <= date_max
    for c in ds.carriers:
        if c['date_min'] is not None and c['date_max'] is not None:
            if c['date_min'] > c['date_max']:
                errors.append(f'Carrier {c["id"]}: date_min ({c["date_min"]}) > date_max ({c["date_max"]})')
        chk_sources('Carrier', c['id'], c['source_ids'])

    # CarrierTraitMix: refs + fraction sums per (carrier, year, domain)
    sums: dict[tuple, float] = {}
    for m in ds.carrier_trait_mix:
        if m['carrier_id'] not in carrier_ids:
            errors.append(f'CarrierTraitMix {m["id"]}: unknown carrier {m["carrier_id"]!r}')
        if m['trait_id'] not in trait_ids:
            errors.append(f'CarrierTraitMix {m["id"]}: unknown trait {m["trait_id"]!r}')
        if m['domain'] not in DOMAINS:
            errors.append(f'CarrierTraitMix {m["id"]}: invalid domain {m["domain"]!r}')
        if m['fraction'] is None or not (0 <= m['fraction'] <= 1):
            errors.append(f'CarrierTraitMix {m["id"]}: fraction {m["fraction"]} not in [0, 1]')
        else:
            key = (m['carrier_id'], m['as_of_year'], m['domain'])
            sums[key] = sums.get(key, 0.0) + m['fraction']
        chk_sources('CarrierTraitMix', m['id'], m['source_ids'])
    for key, total in sums.items():
        if not (0.95 <= total <= 1.05):
            errors.append(f'CarrierTraitMix sum at {key} = {total:.3f}, expected ~1.0')

    # TraitObservations
    for o in ds.trait_observations:
        if o['carrier_id'] and o['carrier_id'] not in carrier_ids:
            errors.append(f'TraitObservation {o["id"]}: unknown carrier {o["carrier_id"]!r}')
        if o['trait_id'] and o['trait_id'] not in trait_ids:
            errors.append(f'TraitObservation {o["id"]}: unknown trait {o["trait_id"]!r}')
        chk_sources('TraitObservation', o['id'], o['source_ids'])

    # PhysicalFeatures + PaleoclimateStates
    for f in ds.physical_features:
        chk_sources('PhysicalFeature', f['id'], f['source_ids'])
    for p in ds.paleoclimate_states:
        chk_sources('PaleoclimateState', p['id'], p['source_ids'])

    # PropagationEvents
    for e in ds.propagation_events:
        for tid in e['source_trait_ids']:
            if tid not in trait_ids:
                errors.append(f'PropagationEvent {e["id"]}: unknown source trait {tid!r}')
        chk_sources('PropagationEvent', e['id'], e['source_ids'])

    # Claims
    for c in ds.claims:
        chk_sources('Claim ' + (c['id'] or '?') + ' supporting', c['id'] or '?', c['supporting_ids'])
        chk_sources('Claim ' + (c['id'] or '?') + ' disputing', c['id'] or '?', c['disputing_ids'])

    # Perspectives
    for p in ds.perspectives:
        if p['parent_id'] and p['parent_id'] not in persp_ids:
            errors.append(f'Perspective {p["id"]}: unknown parent {p["parent_id"]!r}')
        for d in p['domain_scope']:
            if d not in DOMAINS:
                errors.append(f'Perspective {p["id"]}: invalid domain in scope: {d!r}')

    # PerspectiveEndorsements: perspective must exist; if subject_type=Claim, claim must exist
    for e in ds.perspective_endorsements:
        if e['perspective_id'] not in persp_ids:
            errors.append(f'Endorsement {e["id"]}: unknown perspective {e["perspective_id"]!r}')
        if e['subject_type'] == 'Claim' and e['subject_id'] not in claim_ids:
            errors.append(f'Endorsement {e["id"]}: unknown claim {e["subject_id"]!r}')
        if e['stance'] not in {'endorses', 'rejects', 'nuances', 'asserts'}:
            errors.append(f'Endorsement {e["id"]}: invalid stance {e["stance"]!r}')

    return errors


# ---------------------------------------------------------------------------
# Postgres upsert
# ---------------------------------------------------------------------------

UPSERTS = {
    'source': """
        INSERT INTO source (id, type, citation, authors, year, pages_section,
                            doi_url, edition, default_weight, superseded_by, tags)
        VALUES (%(id)s, %(type)s, %(citation)s, %(authors)s, %(year)s, %(pages_section)s,
                %(doi_url)s, %(edition)s, %(default_weight)s, %(superseded_by)s, %(tags)s)
        ON CONFLICT (id) DO UPDATE SET
            type = EXCLUDED.type, citation = EXCLUDED.citation,
            authors = EXCLUDED.authors, year = EXCLUDED.year,
            pages_section = EXCLUDED.pages_section, doi_url = EXCLUDED.doi_url,
            edition = EXCLUDED.edition, default_weight = EXCLUDED.default_weight,
            superseded_by = EXCLUDED.superseded_by, tags = EXCLUDED.tags;
    """,
    'geography': """
        INSERT INTO geo_region (id, display_name, type, centroid, parent_id)
        VALUES (%(id)s, %(display_name)s, %(type)s,
                ST_GeogFromText(%(centroid)s), %(parent_id)s)
        ON CONFLICT (id) DO UPDATE SET
            display_name = EXCLUDED.display_name, type = EXCLUDED.type,
            centroid = EXCLUDED.centroid, parent_id = EXCLUDED.parent_id;
    """,
    'trait': """
        INSERT INTO trait (id, domain, display_name, aliases, first_detected_year,
                           last_detected_year, defining_exemplars, description,
                           politically_sensitive, origin_point)
        VALUES (%(id)s, %(domain)s, %(display_name)s, %(aliases)s, %(first_year)s,
                %(last_year)s, %(exemplars)s, %(description)s,
                %(politically_sensitive)s, ST_GeogFromText(%(origin_point)s))
        ON CONFLICT (id) DO UPDATE SET
            domain = EXCLUDED.domain, display_name = EXCLUDED.display_name,
            aliases = EXCLUDED.aliases, first_detected_year = EXCLUDED.first_detected_year,
            last_detected_year = EXCLUDED.last_detected_year,
            defining_exemplars = EXCLUDED.defining_exemplars,
            description = EXCLUDED.description,
            politically_sensitive = EXCLUDED.politically_sensitive,
            origin_point = EXCLUDED.origin_point, updated_at = now();
    """,
    'carrier': """
        INSERT INTO carrier (id, display_name, type, date_min_year, date_max_year,
                             centroid, archaeological_culture, linguistic_affiliation)
        VALUES (%(id)s, %(display_name)s, %(type)s, %(date_min)s, %(date_max)s,
                ST_GeogFromText(%(centroid)s), %(arch_culture)s, %(linguistic)s)
        ON CONFLICT (id) DO UPDATE SET
            display_name = EXCLUDED.display_name, type = EXCLUDED.type,
            date_min_year = EXCLUDED.date_min_year,
            date_max_year = EXCLUDED.date_max_year,
            centroid = EXCLUDED.centroid,
            archaeological_culture = EXCLUDED.archaeological_culture,
            linguistic_affiliation = EXCLUDED.linguistic_affiliation;
    """,
    'perspective': """
        INSERT INTO perspective (id, display_name, domain_scope, summary, proponents,
                                 methodology_notes, parent_perspective_id,
                                 default_active, status)
        VALUES (%(id)s, %(display_name)s, %(domain_scope)s::trait_domain[],
                %(summary)s, %(proponents)s, %(methodology_notes)s,
                %(parent_id)s, %(default_active)s, %(status)s)
        ON CONFLICT (id) DO UPDATE SET
            display_name = EXCLUDED.display_name,
            domain_scope = EXCLUDED.domain_scope,
            summary = EXCLUDED.summary, proponents = EXCLUDED.proponents,
            methodology_notes = EXCLUDED.methodology_notes,
            parent_perspective_id = EXCLUDED.parent_perspective_id,
            default_active = EXCLUDED.default_active, status = EXCLUDED.status;
    """,
}


def upsert(dsn: str, ds: Dataset, dry_run: bool = False) -> None:
    if dry_run:
        print('[dry-run] would write:')
        for name, rows in [
            ('sources', ds.sources), ('geographies', ds.geographies),
            ('traits', ds.traits), ('trait_relations', ds.trait_relations),
            ('carriers', ds.carriers), ('carrier_trait_mix', ds.carrier_trait_mix),
            ('trait_observations', ds.trait_observations),
            ('physical_features', ds.physical_features),
            ('paleoclimate_states', ds.paleoclimate_states),
            ('propagation_events', ds.propagation_events),
            ('claims', ds.claims), ('perspectives', ds.perspectives),
            ('perspective_endorsements', ds.perspective_endorsements),
        ]:
            print(f'  {name:<28} {len(rows):>4} rows')
        return

    import psycopg  # noqa: lazy import — only needed for real writes

    with psycopg.connect(dsn, autocommit=False) as conn:
        with conn.cursor() as cur:
            # Sources: two-pass to handle self-referential superseded_by FK.
            # Pass 1: insert all sources with superseded_by=NULL.
            for s in ds.sources:
                cur.execute(UPSERTS['source'], {**s, 'superseded_by': None})
            # Pass 2: update superseded_by where set.
            for s in ds.sources:
                if s['superseded_by']:
                    cur.execute(
                        "UPDATE source SET superseded_by = %(superseded_by)s WHERE id = %(id)s",
                        {'id': s['id'], 'superseded_by': s['superseded_by']},
                    )

            # Geographies: two-pass for self-referential parent_id FK.
            # Pass 1: insert without parent_id.
            for g in ds.geographies:
                cur.execute(UPSERTS['geography'], {
                    **g,
                    'centroid': _point_wkt(g['lat'], g['lon']),
                    'parent_id': None,
                })
            # Pass 2: update parent_id where set.
            for g in ds.geographies:
                if g['parent_id']:
                    cur.execute(
                        "UPDATE geo_region SET parent_id = %(parent_id)s WHERE id = %(id)s",
                        {'id': g['id'], 'parent_id': g['parent_id']},
                    )

            # Traits
            for t in ds.traits:
                cur.execute(UPSERTS['trait'], {
                    **t, 'origin_point': _point_wkt(t['lat'], t['lon']),
                })

            # TraitRelations: delete all then re-insert for idempotency
            cur.execute('DELETE FROM trait_relation')
            for r in ds.trait_relations:
                cur.execute("""
                    INSERT INTO trait_relation (child_id, parent_id, relation_type, weight)
                    VALUES (%(child_id)s, %(parent_id)s, %(relation_type)s, %(weight)s);
                """, r)

            # Carriers
            for c in ds.carriers:
                cur.execute(UPSERTS['carrier'], {
                    **c, 'centroid': _point_wkt(c['lat'], c['lon']),
                })

            # CarrierTraitMix: delete all then re-insert for idempotency
            cur.execute('DELETE FROM carrier_trait_mix')
            for m in ds.carrier_trait_mix:
                cur.execute("""
                    INSERT INTO carrier_trait_mix
                        (carrier_id, as_of_year, domain, trait_id, fraction, stderr)
                    VALUES (%(carrier_id)s, %(as_of_year)s, %(domain)s,
                            %(trait_id)s, %(fraction)s, %(stderr)s);
                """, m)

            # TraitObservations
            for o in ds.trait_observations:
                cur.execute("""
                    INSERT INTO trait_observation
                        (id, carrier_id, sample_label, date_min_year, date_max_year,
                         location, domain, trait_id, fraction, stderr, method)
                    VALUES (%(id)s, %(carrier_id)s, %(sample)s, %(date_min)s, %(date_max)s,
                            ST_GeogFromText(%(point)s), %(domain)s, %(trait_id)s,
                            %(fraction)s, %(stderr)s, %(method)s)
                    ON CONFLICT (id) DO UPDATE SET
                        carrier_id = EXCLUDED.carrier_id,
                        sample_label = EXCLUDED.sample_label,
                        date_min_year = EXCLUDED.date_min_year,
                        date_max_year = EXCLUDED.date_max_year,
                        location = EXCLUDED.location, domain = EXCLUDED.domain,
                        trait_id = EXCLUDED.trait_id, fraction = EXCLUDED.fraction,
                        stderr = EXCLUDED.stderr, method = EXCLUDED.method;
                """, {**o, 'point': _point_wkt(o['lat'], o['lon'])})

            # PhysicalFeatures + snapshots
            # Delete snapshots first (idempotency), keep feature master records
            cur.execute('DELETE FROM physical_feature_snapshot')
            for f in ds.physical_features:
                cur.execute("""
                    INSERT INTO physical_feature (id, type, display_name, description)
                    VALUES (%(id)s, %(type)s, %(display_name)s, %(geometry_desc)s)
                    ON CONFLICT (id) DO UPDATE SET
                        type = EXCLUDED.type, display_name = EXCLUDED.display_name,
                        description = EXCLUDED.description;
                """, f)
                if f['as_of_year'] is not None:
                    cur.execute("""
                        INSERT INTO physical_feature_snapshot
                            (feature_id, as_of_year, centroid)
                        VALUES (%(id)s, %(as_of_year)s, ST_GeogFromText(%(centroid)s));
                    """, {**f, 'centroid': _point_wkt(f['lat'], f['lon'])})

            # PaleoclimateStates
            for p in ds.paleoclimate_states:
                cur.execute("""
                    INSERT INTO paleoclimate_state
                        (id, year, scope, sea_level_meters, temp_anomaly_c,
                         ice_volume_relative, co2_ppm)
                    VALUES (%(id)s, %(year)s, %(scope)s, %(sea_level)s,
                            %(temp_anomaly)s, %(ice_volume)s, %(co2_ppm)s)
                    ON CONFLICT (id) DO UPDATE SET
                        year = EXCLUDED.year, scope = EXCLUDED.scope,
                        sea_level_meters = EXCLUDED.sea_level_meters,
                        temp_anomaly_c = EXCLUDED.temp_anomaly_c,
                        ice_volume_relative = EXCLUDED.ice_volume_relative,
                        co2_ppm = EXCLUDED.co2_ppm;
                """, p)

            # PropagationEvents
            for e in ds.propagation_events:
                cur.execute("""
                    INSERT INTO propagation_event
                        (id, display_name, domain, date_min_year, date_max_year,
                         source_trait_ids, source_point, destination_point,
                         mechanism, politically_sensitive)
                    VALUES (%(id)s, %(display_name)s, %(domain)s,
                            %(date_min)s, %(date_max)s, %(source_trait_ids)s,
                            ST_GeogFromText(%(src_point)s),
                            ST_GeogFromText(%(dst_point)s),
                            %(mechanism)s, %(politically_sensitive)s)
                    ON CONFLICT (id) DO UPDATE SET
                        display_name = EXCLUDED.display_name,
                        domain = EXCLUDED.domain,
                        date_min_year = EXCLUDED.date_min_year,
                        date_max_year = EXCLUDED.date_max_year,
                        source_trait_ids = EXCLUDED.source_trait_ids,
                        source_point = EXCLUDED.source_point,
                        destination_point = EXCLUDED.destination_point,
                        mechanism = EXCLUDED.mechanism,
                        politically_sensitive = EXCLUDED.politically_sensitive;
                """, {**e,
                      'src_point': _point_wkt(e['source_lat'], e['source_lon']),
                      'dst_point': _point_wkt(e['dest_lat'], e['dest_lon'])})

            # Claims + claim_source linkage
            # Delete and re-insert for idempotency (claim_id FKs elsewhere are nullable)
            cur.execute('DELETE FROM claim_source')
            cur.execute('DELETE FROM claim')
            claim_db_ids: dict[str, int] = {}
            for c in ds.claims:
                cur.execute("""
                    INSERT INTO claim (subject_type, subject_id, statement,
                                       quantitative_value,
                                       default_aggregated_confidence,
                                       politically_sensitive)
                    VALUES (%(subject_type)s, %(subject_id)s, %(statement)s,
                            %(quant_value)s::jsonb, %(confidence)s, %(politically_sensitive)s)
                    RETURNING id;
                """, {**c, 'quant_value': json.dumps({'raw': c['quant_value']})
                                          if c['quant_value'] else None})
                db_id = cur.fetchone()[0]
                if c['id']:
                    claim_db_ids[c['id']] = db_id
                for sid in c['supporting_ids']:
                    cur.execute("""
                        INSERT INTO claim_source (claim_id, source_id, stance)
                        VALUES (%s, %s, 'supports');
                    """, (db_id, sid))
                for sid in c['disputing_ids']:
                    cur.execute("""
                        INSERT INTO claim_source (claim_id, source_id, stance)
                        VALUES (%s, %s, 'disputes');
                    """, (db_id, sid))

            # Perspectives
            for p in ds.perspectives:
                cur.execute(UPSERTS['perspective'], p)

            # PerspectiveEndorsements: delete and re-insert for idempotency
            cur.execute('DELETE FROM perspective_endorsement')
            for e in ds.perspective_endorsements:
                # Translate spreadsheet claim ID -> DB claim ID where applicable
                subject_id = e['subject_id']
                if e['subject_type'] == 'Claim' and subject_id in claim_db_ids:
                    subject_id = str(claim_db_ids[subject_id])
                cur.execute("""
                    INSERT INTO perspective_endorsement
                        (perspective_id, subject_type, subject_id, stance,
                         override_statement, override_quantitative_value,
                         source_weight_overrides)
                    VALUES (%(perspective_id)s, %(subject_type)s, %(subject_id)s,
                            %(stance)s, %(override_statement)s,
                            %(override_quant_json)s::jsonb, %(source_overrides_json)s::jsonb);
                """, {
                    **e,
                    'subject_id': subject_id,
                    'override_quant_json': json.dumps({'raw': e['override_quant']})
                                           if e['override_quant'] else None,
                    'source_overrides_json': json.dumps(
                        {k.strip(): float(v) for kv in e['source_weight_overrides'].split(';')
                         for k, _, v in [kv.partition(':')] if k.strip() and v.strip()}
                    ) if e['source_weight_overrides'] else None,
                })

        conn.commit()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx', help='Path to template_v0.3.xlsx')
    ap.add_argument('--dsn', help='Postgres DSN (omit for --dry-run)')
    ap.add_argument('--dry-run', action='store_true',
                    help='Validate and print row counts only')
    args = ap.parse_args()

    print(f'Loading {args.xlsx}...')
    ds = load(args.xlsx)

    print('Validating referential integrity...')
    errors = validate(ds)
    if errors:
        print(f'\nVALIDATION FAILED ({len(errors)} errors):', file=sys.stderr)
        for e in errors:
            print(f'  - {e}', file=sys.stderr)
        return 1
    print('OK.')

    if args.dry_run or not args.dsn:
        upsert('', ds, dry_run=True)
        if not args.dsn:
            print('\n(no --dsn provided; dry-run only)')
        return 0

    print(f'Upserting to {args.dsn}...')
    upsert(args.dsn, ds, dry_run=False)
    print('Done.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
